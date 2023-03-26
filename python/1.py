# -*- coding: utf-8 -*-
import numpy as np
import pandas as pd
import openpyxl

# openpyxl.load_workbook(需要打开的excel文件路径)
wb = openpyxl.load_workbook('F:\下载\浏览器下载\一件产品代发明细表2022-3-2.xlsx')
print(type(wb))		# 结果: <class 'openpyxl.workbook.workbook.Workbook'>

sheets_names = wb.sheetnames
print(sheets_names)     # 结果: ['表1', '表2']


sheet = wb.active
print("hfjsdah",sheet)
row_num = sheet.max_row     # 获取当前表中最大的行数
print("row:",row_num)

a1 = sheet['A1']      # A1 表示A列中的第一行，这儿的列号采用的是从A开始的
print(a1)
# 获取单元格中的内容
content = a1.value
print(content)      # 结果是: Rank


rows = [14,15,16,17]

link = []
names = []
for i in range(0,len(rows)) :
    cell = sheet.cell(rows[i],11)
    name = sheet.cell(rows[i],3)
    print(cell.value,name.value)
    link.append(cell.value)
    names.append(name.value)
# for row in range(1, row_num+1):
#     cell = sheet.cell(row, 8)
#     #print(cell.value)


for j in range(0,len(rows)):
    rows[j] = rows[j] -1

title = ['序号','产品名字','链接']

link = np.array(link).reshape(-1,1)
print("链接：",link)
print(link.shape)

rows = np.array(rows).reshape(-1,1)
print("rows:",rows)
names = np.array(names).reshape(-1,1)
print("names:",names)
out = np.hstack((rows,names,link))
print("out:",out)


out = pd.DataFrame(out)
#out.to_excel(r'C:\Users\Hi ZXC\Desktop\1.xlsx')
#title = pd.DataFrame(title)
out.to_excel(r'C:\Users\Hi ZXC\Desktop\1.xlsx',header=title,index=None)

