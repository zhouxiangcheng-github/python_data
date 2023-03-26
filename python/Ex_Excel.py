# -*- coding: utf-8 -*-
# by zxc
import pandas as pd
import openpyxl
import numpy as np

wb = openpyxl.load_workbook('F:\下载\浏览器下载\一件产品代发明细表2022-3-2 (1).xlsx')

sheet = wb.active



data = pd.DataFrame(pd.read_excel('F:\下载\浏览器下载\一件产品代发明细表2022-3-2 (1).xlsx',usecols=[0],
                       names=None)).values

data = list(data)

index = []
rows = [6056,6141,6163,6643,9345,9360,7935,7468,7532,8574,9313,9161,9147,9146,8490,8494,8501,8505,8507,8328,8309,8310,8311,8155,8156,8157,8164,8165]
'''
获取元素对应的行号
1.获取元素对应的位置----->data.index(rows[i])
2.获得该位置对应的行号----->data.index(rows[i]) + 2   
#加2是因为行号错开了1，另外索引值是从0开始，因此加2
'''
for i in range(0,len(rows)):
    value = data.index(rows[i]) + 2
    index.append(value)
#print(index)

#测试
# rows = [13,14,15,16]
# rows_new = [13,14,15,16]




link = []
names = []
d_prices = []
prices = []



rows_new = index

for i in range(0,len(rows_new)) :
    cell = sheet.cell(rows_new[i],11)    #读取链接，cell()的参数是---->行列位置----->第rows_new[i]行，11列
    name = sheet.cell(rows_new[i],3)     #读取产品名字
    d_price = sheet.cell(rows_new[i],7)   #读取代发价格
    price = sheet.cell(rows_new[i],8)        #读取价格
    #print(cell.value,name.value)
    link.append(cell.value)
    names.append(name.value)
    d_prices.append(d_price.value)
    prices.append(price.value)





#制定标题
title = ['序号','产品名字','链接','代发价格','价格']
'''
分别将序号，产品名字，链接的列表转换为数组
'''
link = np.array(link).reshape(-1,1)
rows = np.array(rows).reshape(-1,1)
names = np.array(names).reshape(-1,1)
d_prices = np.array(d_prices).reshape(-1,1)
prices = np.array(prices).reshape(-1,1)
#拼接成数组
out = np.hstack((rows,names,link,d_prices,prices))



out = pd.DataFrame(out)

out.to_excel(r'C:\Users\Hi ZXC\Desktop\1.xlsx',header=title,index=None)

